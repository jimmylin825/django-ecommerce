import csv

from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.csrf import csrf_exempt
from .models import Order, OrderItem, Product, Category
from django.utils import timezone
from django.db.models import Sum, F
from django import forms
from datetime import date
from .forms import OrderMessageForm

# Create your views here.


def product_list(request):
    products = Product.objects.filter(is_available = True)
    category = request.GET.get('category')
    query = request.GET.get('q')
    min_price = request.GET.get('min_price')
    max_price = request.GET.get('max_price')
    if category:
        products = products.filter(category=category)
    if query:
        products = products.filter(name__icontains=query)
    if min_price:
        products = products.filter(price__gte=min_price)
    if max_price:
        products = products.filter(price__lte=max_price)

    # 取得所有不同的分類（方便前端顯示篩選）
    categories = Category.objects.all()

    return render(request, 'shop/product_list.html', {'products': products, 'categories': categories})


def product_detail(request, product_id):
    product = get_object_or_404(Product, id = product_id)
    images = product.images.all()
    # 推薦同分類的其他商品（最多 4 個，不包含當前商品）
    related_products = Product.objects.filter(
        category=product.category
    ).exclude(id=product.id)[:4]

    return render(request, 'shop/product_detail.html',
                  {'product' : product,
                   'related_products' : related_products,
                   'images' : images})

def add_to_cart(request, product_id):
    product = get_object_or_404(Product, id = product_id)
    quantity = int(request.POST.get('quantity', 1))

    #初始化購物車
    cart = request.session.get('cart', {})

    #加入或更新購物車數量
    cart[str(product_id)] = cart.get(str(product_id), 0) + quantity

    request.session['cart'] = cart  #寫回session
    return redirect('cart_view')

def cart_view(request):
    cart = request.session.get('cart', {})
    cart_items = []
    total = 0

    for product_id, quantity in cart.items():
        product = get_object_or_404(Product, id = int(product_id))
        subtotal = product.price * quantity
        cart_items.append({
            'product' : product,
            'quantity' : quantity,
            'subtotal' : subtotal
        })
        total += subtotal

    return render(request, 'shop/cart.html', {'cart_items' : cart_items, 'total' : total, 'step' : 1, 'error' : '發生錯誤請稍後再試'})



def update_cart(request, product_id):
    if request.method == 'POST':
        quantity = int(request.POST.get('quantity', 1))
        cart = request.session.get('cart', {})
        cart[str(product_id)] = quantity
        request.session['cart'] = cart
    return redirect('cart_view')


def remove_cart(request, product_id):
    if request.method == 'POST':
        cart = request.session.get('cart', {})
        cart.pop(str(product_id), None)
        request.session['cart'] = cart
    return redirect('cart_view')


def create_order(request):
    if request.method == 'POST':
        cart = request.session.get('cart', {})
        if not cart:
            return HttpResponse('購物車為空，無法建立訂單')
    if request.user.is_authenticated:
        customer = request.user
        customer_name = request.user.username
        customer_email = request.user.email
    else:
        customer = None
        customer_name = '匿名用戶'
        customer_email = ''
    order = Order.objects.create(
        customer = customer,
        customer_email = customer_email,
        customer_name = customer_name,
        created_at = timezone.now()
    )

    for product_id, quantity in cart.items():
        product = get_object_or_404(Product, id = product_id)

        # 檢查庫存是否足夠
        if product.stock < quantity:
            return HttpResponse(f"{product.name} 庫存不足，僅剩 {product.stock} 件。")

        OrderItem.objects.create(
            order=order,
            product=product,
            quantity=quantity,
            price=product.price
        )
        # 扣除庫存
        product.stock -= quantity
        product.save()
        
    #清空購物車
    request.session['cart'] = {}

    return redirect('order_success', order_id = order.id)

def order_success(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    return render(request, 'shop/order_success.html', {'order': order, 'step' : 3, 'error' : '發生錯誤請稍後再試'})

#顯示所有訂單
def order_list(request):
    if request.user.is_authenticated:
        orders = Order.objects.filter(customer = request.user).order_by('created_at')
    else:
        return HttpResponse("請先登入才能查看您的訂單記錄")
    return render(request, 'shop/order_list.html', {'orders' : orders})

#顯示單一訂單詳情
def order_detail(request, order_id):
    order = get_object_or_404(Order, id = order_id)
    # 限制已登入用戶只能看自己的訂單（不可看匿名訂單）
    if request.user.is_authenticated:
        if order.customer != request.user:
            return HttpResponseForbidden("您無權限查看此訂單")
    else:
        # 未登入時，只能看 anonymous 訂單（customer=None）
        if order.customer is not None:
            return HttpResponseForbidden("您無權限查看此訂單")

    items = order.items.all()
    return render(request, 'shop/order_detail.html', {
        'order': order,
        'items': items
    })

class SalesStatsForm(forms.Form):
    start_date = forms.DateField(label = '起始日期', widget=forms.DateInput(attrs={'type' : 'date'}))
    end_date = forms.DateField(label = '結束日期', widget=forms.DateInput(attrs={'type' : 'date'}))


@staff_member_required
def sales_stats(request):
    start_date = request.GET.get('start_date', str(date.today().replace(day=1)))
    end_date = request.GET.get('end_date', str(date.today()))

    orders = Order.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    )

    total_sales = orders.aggregate(total=Sum(F('items__quantity') * F('items__price')))['total'] or 0

    items = (
        OrderItem.objects
        .filter(order__in=orders)
        .values('product__name')
        .annotate(
            total_sold=Sum('quantity'),
            total_amount=Sum(F('quantity') * F('price'))
        )
        .order_by('-total_sold')
    )

    # 匯出 CSV
    if 'export' in request.GET and request.GET['export'] == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="sales_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['商品名稱', '總銷量', '總金額'])
        for item in items:
            writer.writerow([item['product__name'], item['total_sold'], item['total_amount']])
        return response

    # 匯出 Excel
    if 'export' in request.GET and request.GET['export'] == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "銷售統計"
        headers = ['商品名稱', '總銷量', '總金額']
        ws.append(headers)

        for row in items:
            ws.append([row['product__name'], row['total_sold'], row['total_amount']])

        for col in range(1, len(headers) + 1):
            ws[f'{get_column_letter(col)}1'].font = Font(bold=True)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="sales_report.xlsx"'
        wb.save(response)
        return response

    form = SalesStatsForm(initial={
        'start_date': start_date,
        'end_date': end_date,
    })

    return render(request, 'admin/sales_stats.html', {
        'form': form,
        'orders': orders,
        'total_sales': total_sales,
        'best_sellers': items
    })

@login_required
def customer_report(request, order_id):
    order = get_object_or_404(Order, id = order_id)
    if request.method == 'POST':
        form = OrderMessageForm(request.POST)
        if form.is_valid():
            msg = form.save(commit=False)
            msg.order = order
            msg.sender = request.user
            msg.save()
            return redirect('customer_report', order_id=order_id)
    else:
        form = OrderMessageForm()
    items = order.items.all()
    messages = order.messages.all()  # 補上這行
    return render(request, 'shop/customer_report.html', {
        'order': order,
        'items': items,
        'form': form,
        'messages': messages,
        'user': request.user
    })

